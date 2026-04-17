import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart

def load_and_clean_data(file_path):
    """Load and clean the sales data"""
    print("\n" + "=" * 60)
    print("LOADING DATA")
    print("=" * 60)
    
    df = pd.read_excel(file_path)
    df.columns = df.columns.str.strip()
    
    # Convert dates
    df['arrival_date'] = pd.to_datetime(df['arrival_date'])
    df['departure_date'] = pd.to_datetime(df['departure_date'])
    
    # Calculate key metrics
    df['nights'] = (df['departure_date'] - df['arrival_date']).dt.days
    df['nights'] = df['nights'].replace(0, 1)  # Minimum 1 night
    df['revenue'] = df['price'] * df['nights']
    
    # Time periods
    df['month'] = df['arrival_date'].dt.to_period('M')
    df['month_name'] = df['arrival_date'].dt.strftime('%B %Y')
    df['year'] = df['arrival_date'].dt.year
    df['quarter'] = 'Q' + df['arrival_date'].dt.quarter.astype(str)
    
    # Guest counts
    df['adult'] = pd.to_numeric(df['adult'], errors='coerce').fillna(0).astype(int)
    df['(child)'] = pd.to_numeric(df['(child)'], errors='coerce').fillna(0).astype(int)
    df['total_guests'] = df['adult'] + df['(child)']
    
    print(f"✓ Loaded {len(df):,} records")
    print(f"✓ Date range: {df['arrival_date'].min().date()} to {df['arrival_date'].max().date()}")
    print(f"✓ Hotels: {df['hotel_name'].nunique()}")
    print(f"✓ Total revenue: RM {df['revenue'].sum():,.0f}")
    
    return df

def get_monthly_summary(df, target_month=None):
    """Generate monthly summary statistics"""
    
    if target_month:
        # Format: '2024-01' or 'January 2024'
        try:
            target_period = pd.Period(target_month, freq='M')
        except:
            # Try parsing as full month name
            target_period = pd.Period(pd.to_datetime(target_month), freq='M')
        monthly_data = df[df['month'] == target_period].copy()
        month_label = target_period.strftime('%B %Y')
    else:
        # Use latest month
        latest_month = df['month'].max()
        monthly_data = df[df['month'] == latest_month].copy()
        month_label = latest_month.strftime('%B %Y')
    
    print("\n" + "=" * 60)
    print(f"GENERATING REPORT FOR: {month_label}")
    print("=" * 60)
    
    if len(monthly_data) == 0:
        print(f"⚠️  No data found for {month_label}")
        return None, month_label
    
    print(f"Records in this month: {len(monthly_data):,}")
    
    # Overall KPIs
    summary = {
        'Month': month_label,
        'Total Bookings': len(monthly_data),
        'Total Revenue (RM)': monthly_data['revenue'].sum(),
        'Avg Booking Value (RM)': monthly_data['revenue'].mean(),
        'Total Nights Booked': monthly_data['nights'].sum(),
        'Avg Length of Stay': monthly_data['nights'].mean(),
        'Total Guests': int(monthly_data['total_guests'].sum()),
        'Avg Guests per Booking': monthly_data['total_guests'].mean(),
        'Avg Nightly Rate (RM)': monthly_data['price'].mean()
    }
    
    # By Hotel
    by_hotel = monthly_data.groupby('hotel_name').agg({
        'revenue': ['sum', 'mean'],
        'hotel_name': 'count',  # booking count
        'nights': 'sum',
        'total_guests': 'sum',
        'price': 'mean'
    })
    by_hotel.columns = ['Total Revenue', 'Avg Booking Value', 'Bookings', 'Total Nights', 'Total Guests', 'Avg Nightly Rate']
    by_hotel['Revenue %'] = (by_hotel['Total Revenue'] / by_hotel['Total Revenue'].sum() * 100).round(1)
    by_hotel = by_hotel.sort_values('Total Revenue', ascending=False)
    
    # By Channel
    by_channel = monthly_data.groupby('dis_channel').agg({
        'revenue': ['sum', 'mean'],
        'dis_channel': 'count'
    })
    by_channel.columns = ['Total Revenue', 'Avg Booking Value', 'Bookings']
    by_channel['Revenue %'] = (by_channel['Total Revenue'] / by_channel['Total Revenue'].sum() * 100).round(1)
    by_channel = by_channel.sort_values('Total Revenue', ascending=False)
    
    # By Room Type
    by_room = monthly_data.groupby('room_type').agg({
        'revenue': ['sum', 'mean'],
        'room_type': 'count',
        'price': 'mean'
    })
    by_room.columns = ['Total Revenue', 'Avg Booking Value', 'Bookings', 'Avg Nightly Rate']
    by_room['Revenue %'] = (by_room['Total Revenue'] / by_room['Total Revenue'].sum() * 100).round(1)
    by_room = by_room.sort_values('Total Revenue', ascending=False)
    
    # By Customer Segment
    by_segment = monthly_data.groupby('cus_seg').agg({
        'revenue': ['sum', 'mean'],
        'cus_seg': 'count'
    })
    by_segment.columns = ['Total Revenue', 'Avg Booking Value', 'Bookings']
    by_segment['Revenue %'] = (by_segment['Total Revenue'] / by_segment['Total Revenue'].sum() * 100).round(1)
    by_segment = by_segment.sort_values('Total Revenue', ascending=False)
    
    # By Membership
    by_membership = monthly_data.groupby('membership').agg({
        'revenue': ['sum', 'mean'],
        'membership': 'count'
    })
    by_membership.columns = ['Total Revenue', 'Avg Booking Value', 'Bookings']
    by_membership['Revenue %'] = (by_membership['Total Revenue'] / by_membership['Total Revenue'].sum() * 100).round(1)
    by_membership = by_membership.sort_values('Total Revenue', ascending=False)
    
    # By Payment Method
    by_payment = monthly_data.groupby('payment_method').agg({
        'revenue': 'sum',
        'payment_method': 'count'
    })
    by_payment.columns = ['Total Revenue', 'Bookings']
    by_payment['Revenue %'] = (by_payment['Total Revenue'] / by_payment['Total Revenue'].sum() * 100).round(1)
    by_payment = by_payment.sort_values('Total Revenue', ascending=False)
    
    # By Meal Plan
    by_meal = monthly_data.groupby('meal_plan').agg({
        'revenue': 'sum',
        'meal_plan': 'count'
    })
    by_meal.columns = ['Total Revenue', 'Bookings']
    by_meal['Revenue %'] = (by_meal['Total Revenue'] / by_meal['Total Revenue'].sum() * 100).round(1)
    by_meal = by_meal.sort_values('Total Revenue', ascending=False)
    
    return {
        'summary': summary,
        'by_hotel': by_hotel,
        'by_channel': by_channel,
        'by_room': by_room,
        'by_segment': by_segment,
        'by_membership': by_membership,
        'by_payment': by_payment,
        'by_meal': by_meal,
        'raw_data': monthly_data
    }, month_label

def create_excel_report(report_data, month_label, output_path):
    """Create formatted Excel report"""
    print("\n" + "=" * 60)
    print("CREATING EXCEL REPORT")
    print("=" * 60)
    
    summary = report_data['summary']
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # === SUMMARY SHEET ===
        summary_df = pd.DataFrame.from_dict(summary, orient='index', columns=['Value'])
        summary_df.to_excel(writer, sheet_name='Executive Summary')
        
        # === BY HOTEL SHEET ===
        report_data['by_hotel'].to_excel(writer, sheet_name='By Hotel')
        
        # === BY CHANNEL SHEET ===
        report_data['by_channel'].to_excel(writer, sheet_name='By Channel')
        
        # === BY ROOM TYPE SHEET ===
        report_data['by_room'].to_excel(writer, sheet_name='By Room Type')
        
        # === BY SEGMENT SHEET ===
        report_data['by_segment'].to_excel(writer, sheet_name='By Customer Segment')
        
        # === BY MEMBERSHIP SHEET ===
        report_data['by_membership'].to_excel(writer, sheet_name='By Membership')
        
        # === BY PAYMENT METHOD SHEET ===
        report_data['by_payment'].to_excel(writer, sheet_name='By Payment Method')
        
        # === BY MEAL PLAN SHEET ===
        report_data['by_meal'].to_excel(writer, sheet_name='By Meal Plan')
        
        # === RAW DATA SHEET ===
        report_data['raw_data'].to_excel(writer, sheet_name='Raw Data', index=False)
    
    # Format the workbook
    wb = openpyxl.load_workbook(output_path)
    
    # Format Executive Summary sheet
    ws = wb['Executive Summary']
    ws.insert_rows(1)
    ws['A1'] = f'Monthly Sales Report - {month_label}'
    ws['A1'].font = Font(size=16, bold=True, color='1D428A')
    ws.merge_cells('A1:B1')
    
    # Header styling for all sheets
    header_fill = PatternFill(start_color='1D428A', end_color='1D428A', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Style headers
        for cell in ws[1 if sheet_name == 'Raw Data' else 2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"✓ Report saved to: {output_path}")

def generate_insights(report_data):
    """Generate automated insights from the data"""
    print("\n" + "=" * 60)
    print("KEY INSIGHTS")
    print("=" * 60)
    
    summary = report_data['summary']
    by_hotel = report_data['by_hotel']
    by_channel = report_data['by_channel']
    by_room = report_data['by_room']
    
    # Insight 1: Top performing hotel
    top_hotel = by_hotel.index[0]
    top_hotel_rev = by_hotel.iloc[0]['Total Revenue']
    top_hotel_pct = by_hotel.iloc[0]['Revenue %']
    
    print(f"\n1.Top Performing Hotel: {top_hotel}")
    print(f"   Revenue: RM {top_hotel_rev:,.0f} ({top_hotel_pct:.1f}% of total)")
    print(f"   Bookings: {by_hotel.iloc[0]['Bookings']:,.0f}")
    print(f"   Avg Booking Value: RM {by_hotel.iloc[0]['Avg Booking Value']:,.0f}")
    
    # Insight 2: Top channel
    top_channel = by_channel.index[0]
    top_channel_pct = by_channel.iloc[0]['Revenue %']
    
    print(f"\n2.Leading Distribution Channel: {top_channel}")
    print(f"   Contribution: {top_channel_pct:.1f}% of revenue")
    print(f"   Bookings: {by_channel.iloc[0]['Bookings']:,.0f}")
    print(f"   Avg Booking Value: RM {by_channel.iloc[0]['Avg Booking Value']:,.0f}")
    
    # Insight 3: Most popular room type
    top_room = by_room.index[0]
    top_room_pct = by_room.iloc[0]['Revenue %']
    
    print(f"\n3.Most Popular Room Type: {top_room}")
    print(f"   Revenue Contribution: {top_room_pct:.1f}%")
    print(f"   Bookings: {by_room.iloc[0]['Bookings']:,.0f}")
    print(f"   Avg Nightly Rate: RM {by_room.iloc[0]['Avg Nightly Rate']:,.0f}")
    
    # Insight 4: Overall performance metrics
    print(f"\n4. 📈 Overall Performance:")
    print(f"   Total Revenue: RM {summary['Total Revenue (RM)']:,.0f}")
    print(f"   Total Bookings: {summary['Total Bookings']:,}")
    print(f"   Avg Booking Value: RM {summary['Avg Booking Value (RM)']:,.0f}")
    print(f"   Avg Length of Stay: {summary['Avg Length of Stay']:.1f} nights")
    print(f"   Avg Party Size: {summary['Avg Guests per Booking']:.1f} guests")
    
    print("\n" + "=" * 60)

def main():
    """Main execution function"""
    print("\n" + "=" * 60)
    print("HOTEL GROUP MONTHLY SALES REPORT GENERATOR")
    print("=" * 60)
    
    try:
        # Load data
        df = load_and_clean_data('data/sample_data.xlsx')
        
        # Generate report for latest month
        # To specify a month: target_month='2024-06' or 'June 2024'
        report_data, month_label = get_monthly_summary(df, target_month=None)
        
        if report_data is None:
            print("\nCannot generate report. Exiting.")
            return
        
        # Create output directory
        output_dir = Path('output')
        output_dir.mkdir(exist_ok=True)
        
        # Generate filename
        month_str = month_label.replace(' ', '_')
        output_path = output_dir / f'Monthly_Sales_Report_{month_str}.xlsx'
        
        # Create Excel report
        create_excel_report(report_data, month_label, output_path)
        
        # Generate insights
        generate_insights(report_data)
        
        print("\n" + "=" * 60)
        print("REPORT GENERATION COMPLETE!")
        print("=" * 60)
        print(f"\nOutput file: {output_path}")
        print(f"Report covers: {month_label}")
        print(f"Total revenue: RM {report_data['summary']['Total Revenue (RM)']:,.0f}")
        print(f"Total bookings: {report_data['summary']['Total Bookings']:,}")
        
    except FileNotFoundError:
        print("\n Error: data/sample_data.xlsx not found!")
        print("   Make sure your data file is in the correct location.")
    except Exception as e:
        print(f"\n Error occurred: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()